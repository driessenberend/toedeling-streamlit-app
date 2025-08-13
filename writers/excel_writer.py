from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Optional, Any
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from config import SHEET_CODEMAP, TARGET_COLUMNS
from loaders.customer_workbook import (
    read_header,
    ensure_target_columns,
    iter_data_rows,
    write_results,
)
from loaders.schema_loader import load_codeschema_excel, CodeRule
from logic.prompts import build_system_prompt, build_user_prompt
from logic.classifier import (
    rank_candidates,
    build_row_text,
    pick_code_with_llm,
    simple_rules_fallback,
)
from llm_providers.base import LLMClient
from llm_providers.openai_provider import OpenAIClient


def _select_provider(name: str) -> LLMClient:
    """
    Kies en instantieer de LLM-provider (modulair). Voeg hier eenvoudig extra providers toe.
    """
    if name == "openai":
        return OpenAIClient()
    raise RuntimeError(f"Onbekende provider: {name}")


def process_workbook(
    *,
    customer_file,
    schema_file,
    header_rows: Dict[str, int],
    provider_name: str,
    model: str,
    temperature: float,
    top_k_codes: int,
    dry_run: bool,
    language: str,
) -> BytesIO:
    """
    Verwerkt het klantbestand:
    - Leest codeschema (formatie/kosten/opbrengsten)
    - Loopt door relevante 'Oplegger'-tabbladen
    - Schrijft 'Codering AI', 'Argumentatie AI', 'Opmerkingen/aannames vanuit Berenschot'
    - Retourneert een BytesIO met het aangepaste workbook
    """
    wb: Workbook = load_workbook(customer_file)
    schema = load_codeschema_excel(schema_file)

    # Provider (modulair)
    llm: Optional[LLMClient] = None
    if not dry_run:
        try:
            llm = _select_provider(provider_name)
        except Exception:
            # Val veilig terug op offline modus als de provider faalt (bijv. geen API-sleutel)
            llm = None

    system_prompt = build_system_prompt(language=language)

    # Voor snelle lookup van target-kolomtitels (om ze uit de context te filteren)
    target_titles_lc = {v.strip().lower() for v in TARGET_COLUMNS.values()}

    for ws in wb.worksheets:
        name_l = ws.title.strip().lower()
        match_key: Optional[str] = None
        for key in SHEET_CODEMAP:
            if key in name_l:
                match_key = key
                break
        if not match_key:
            continue  # Irrelevant tabblad

        category = SHEET_CODEMAP[match_key]  # 'formatie' | 'kosten' | 'opbrengsten'
        rules: List[CodeRule] = schema.get(category, [])
        if not rules:
            # Geen regels beschikbaar voor deze categorie -> sla over
            continue

        # Header-rij bepalen (met fallback op 1)
        header_row = header_rows.get(name_l, header_rows.get(match_key, 1))

        # Header inlezen en target-kolommen garanderen
        header_map = read_header(ws, header_row)
        target_indices = ensure_target_columns(ws, header_row)

        # Contextkolommen = alle headers behalve de 3 doelkolommen
        context_cols: Dict[str, int] = {
            col_name: idx
            for col_name, idx in header_map.items()
            if col_name.strip().lower() not in target_titles_lc
        }

        for r in iter_data_rows(ws, header_row):
            # Bouw context uit de rij
            context: Dict[str, Any] = {
                k: ws.cell(row=r, column=idx).value for k, idx in context_cols.items()
            }

            # Kandidaten shortlist via fuzzy matching
            text_for_rank = build_row_text(context)
            candidates = rank_candidates(text_for_rank, rules, top_k=top_k_codes) if rules else []

            # Kies code via LLM of via eenvoudige fallback
            if llm is None:
                result = simple_rules_fallback(context, candidates if candidates else rules)
            else:
                user_prompt = build_user_prompt(context, candidates if candidates else rules, category=category)
                try:
                    result = pick_code_with_llm(
                        llm,
                        model=model,
                        system_prompt=system_prompt,
                        user_prompt=user_prompt,
                        temperature=temperature,
                    )
                except Exception:
                    # Robuust: bij fout terugvallen op heuristiek
                    result = simple_rules_fallback(context, candidates if candidates else rules)

            # Schrijf resultaat in de juiste kolommen
            write_results(
                ws,
                row_idx=r,
                target_indices=target_indices,
                code=result.code,
                argument=result.argumentatie,
                note=result.vraag,
            )

    # Schrijf terug naar bytes
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
