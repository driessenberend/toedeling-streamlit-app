from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Optional, Any
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from config import TARGET_COLUMNS

@dataclass
class SheetSpec:
    name: str
    header_row: int

def read_header(ws: Worksheet, header_row: int) -> Dict[str, int]:
    """Return dict: header_name -> column_index (1-based)."""
    headers: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        val = cell.value
        if val is None:
            continue
        name = str(val).strip()
        if name:
            headers[name] = col_idx
    return headers

def ensure_target_columns(ws: Worksheet, header_row: int) -> Dict[str, int]:
    headers = read_header(ws, header_row)
    max_col = ws.max_column
    indices: Dict[str, int] = {}
    for key, title in TARGET_COLUMNS.items():
        found_idx = None
        for hname, idx in headers.items():
            if hname.lower() == title.lower():
                found_idx = idx
                break
        if found_idx is None:
            max_col += 1
            ws.cell(row=header_row, column=max_col, value=title)
            found_idx = max_col
        indices[key] = found_idx
    return indices

def iter_data_rows(ws: Worksheet, header_row: int):
    """Yield 1-based row indices for non-empty data rows after the header."""
    start = header_row + 1
    for r in range(start, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v in (None, "") for v in row_vals):
            continue
        yield r

def build_row_context(ws: Worksheet, header_map: Dict[str, int], row_idx: int) -> Dict[str, Any]:
    ctx = {}
    for name, col_idx in header_map.items():
        ctx[name] = ws.cell(row=row_idx, column=col_idx).value
    return ctx

def write_results(ws: Worksheet, row_idx: int, target_indices: Dict[str, int],
                  code: Optional[str], argument: Optional[str], note: Optional[str]) -> None:
    if code is not None:
        ws.cell(row=row_idx, column=target_indices["codering_ai"], value=code)
    if argument is not None:
        ws.cell(row=row_idx, column=target_indices["argumentatie_ai"], value=argument)
    if note is not None:
        ws.cell(row=row_idx, column=target_indices["opmerkingen"], value=note)
